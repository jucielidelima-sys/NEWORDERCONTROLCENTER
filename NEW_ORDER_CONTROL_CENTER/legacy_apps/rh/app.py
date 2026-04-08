from datetime import date, datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import plotly.graph_objects as go
from openpyxl import load_workbook


ARQUIVO = "QUADRO COLABORADORES.xlsx"
ABA_BASE = "RELAÇÃO DE COLABORADORES 19"
CAB_ABS = ["Data", "Cadastro", "Nome", "Admissão", "Cargo", "Setor", "Motivo", "Observação", "Lançado em"]
CAB_TURN = ["Data", "Cadastro", "Nome", "Admissão", "Cargo", "Setor", "Motivo", "Observação", "Lançado em"]
CAB_NOV = ["Data Registro", "Cadastro", "Nome", "Admissão", "Cargo", "Setor", "Origem", "Observação", "Lançado em"]

st.set_page_config(page_title="RH • Pessoas", layout="wide")

st.markdown(
    '''
    <style>
    .stApp{
        background:
            radial-gradient(circle at top left, rgba(0,195,255,0.10), transparent 18%),
            radial-gradient(circle at top right, rgba(255,120,0,0.08), transparent 18%),
            linear-gradient(135deg, #0a1017 0%, #101924 45%, #0c1219 100%);
        color:#edf3f9;
    }
    .rh-card{
        border-radius:20px;
        padding:18px;
        background:linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.025));
        border:1px solid rgba(255,255,255,0.08);
        box-shadow:0 10px 24px rgba(0,0,0,0.22);
    }
    .rh-lbl{
        color:#9fb0c2;
        font-size:12px;
        text-transform:uppercase;
        font-weight:800;
        letter-spacing:1px;
    }
    .rh-val{
        color:#ffffff;
        font-size:34px;
        font-weight:950;
        margin-top:8px;
    }
    .rh-sub{
        color:#91a1b3;
        font-size:12px;
        margin-top:8px;
    }
    </style>
    ''',
    unsafe_allow_html=True
)

def caminho_arquivo() -> Path:
    return Path(__file__).parent / ARQUIVO

def carregar_base() -> pd.DataFrame:
    path = caminho_arquivo()
    df = pd.read_excel(path, sheet_name=ABA_BASE, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    cols = ["Cadastro", "Nome", "Admissão", "Cargo", "SETOR"]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df = df[cols].copy()
    df["Nome"] = df["Nome"].astype(str).str.strip()
    df["SETOR"] = df["SETOR"].astype(str).str.strip()
    df["Admissão"] = pd.to_datetime(df["Admissão"], errors="coerce", dayfirst=True)
    return df

def garantir_abas():
    path = caminho_arquivo()
    wb = load_workbook(path)
    estrutura = {
        "ABSENTEISMO": CAB_ABS,
        "TURNOVER": CAB_TURN,
        "NOVAS_CONTRATACOES": CAB_NOV,
    }
    changed = False
    for aba, headers in estrutura.items():
        if aba not in wb.sheetnames:
            ws = wb.create_sheet(aba)
            ws.append(headers)
            changed = True
        else:
            ws = wb[aba]
            if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
                ws.delete_rows(1, 1)
                ws.append(headers)
                changed = True
    if changed:
        wb.save(path)

def ler_aba(nome_aba: str, headers: list) -> pd.DataFrame:
    path = caminho_arquivo()
    wb = load_workbook(path, data_only=True)
    if nome_aba not in wb.sheetnames:
        return pd.DataFrame(columns=headers)
    ws = wb[nome_aba]
    values = list(ws.values)
    if not values:
        return pd.DataFrame(columns=headers)
    df = pd.DataFrame(values[1:], columns=values[0])
    df = df.fillna("")
    return df

def append_row(nome_aba: str, row: list):
    path = caminho_arquivo()
    wb = load_workbook(path)
    ws = wb[nome_aba]
    ws.append(row)
    wb.save(path)

def info_colaborador(df_base: pd.DataFrame, nome: str) -> dict:
    rec = df_base[df_base["Nome"] == nome]
    if rec.empty:
        return {"Cadastro": "", "Nome": nome, "Admissão": "", "Cargo": "", "SETOR": ""}
    r = rec.iloc[0]
    adm = r.get("Admissão", "")
    if pd.notna(adm):
        adm = pd.Timestamp(adm).strftime("%d/%m/%Y")
    else:
        adm = ""
    return {
        "Cadastro": r.get("Cadastro", ""),
        "Nome": r.get("Nome", ""),
        "Admissão": adm,
        "Cargo": r.get("Cargo", ""),
        "SETOR": r.get("SETOR", ""),
    }

def preparar_datas(df: pd.DataFrame, col: str) -> pd.DataFrame:
    out = df.copy()
    if col in out.columns:
        out[col] = pd.to_datetime(out[col], errors="coerce", dayfirst=True)
        out["AnoMes"] = out[col].dt.to_period("M").astype(str)
        out["MesLabel"] = out[col].dt.strftime("%m/%Y")
    else:
        out["AnoMes"] = ""
        out["MesLabel"] = ""
    return out

def contar_mes(df, col, ref):
    if df.empty or col not in df.columns:
        return 0
    s = pd.to_datetime(df[col], errors="coerce", dayfirst=True)
    return int(((s.dt.month == ref.month) & (s.dt.year == ref.year)).sum())

def kpi_card(label, value, sub):
    st.markdown(
        f'''
        <div class="rh-card">
            <div class="rh-lbl">{label}</div>
            <div class="rh-val">{value}</div>
            <div class="rh-sub">{sub}</div>
        </div>
        ''',
        unsafe_allow_html=True
    )

def plot_bar(df, x, y, title):
    fig = go.Figure()
    fig.add_trace(go.Bar(x=df[x], y=df[y], name=title))
    fig.update_layout(
        title=title,
        height=340,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(255,255,255,0.02)",
        font=dict(color="#edf3f9"),
        margin=dict(l=20, r=20, t=50, b=20),
    )
    fig.update_xaxes(showgrid=False)
    fig.update_yaxes(gridcolor="rgba(255,255,255,0.08)")
    st.plotly_chart(fig, use_container_width=True)

garantir_abas()
df_base = carregar_base()
nomes = sorted([n for n in df_base["Nome"].dropna().unique().tolist() if str(n).strip()])

abs_df = preparar_datas(ler_aba("ABSENTEISMO", CAB_ABS), "Data")
turn_df = preparar_datas(ler_aba("TURNOVER", CAB_TURN), "Data")
nov_df = preparar_datas(ler_aba("NOVAS_CONTRATACOES", CAB_NOV), "Data Registro")

hoje = pd.Timestamp.today()
abs_mes = contar_mes(abs_df, "Data", hoje)
turn_mes = contar_mes(turn_df, "Data", hoje)
adm_mes = contar_mes(nov_df, "Data Registro", hoje)
base_total = len(df_base)

taxa_abs = (abs_mes / base_total * 100) if base_total else 0
taxa_turn = (turn_mes / base_total * 100) if base_total else 0

setores = sorted([s for s in df_base["SETOR"].dropna().unique().tolist() if str(s).strip()])
opcoes_setor = ["Todos"] + setores
setor_sel = st.sidebar.selectbox("Filtrar setor", opcoes_setor, index=0)

meses_disp = sorted(set([m for m in abs_df.get("AnoMes", pd.Series(dtype=str)).astype(str).tolist() if m and m != "NaT"] +
                        [m for m in turn_df.get("AnoMes", pd.Series(dtype=str)).astype(str).tolist() if m and m != "NaT"] +
                        [m for m in nov_df.get("AnoMes", pd.Series(dtype=str)).astype(str).tolist() if m and m != "NaT"]))
opcoes_mes = ["Todos"] + meses_disp
mes_sel = st.sidebar.selectbox("Filtrar mês", opcoes_mes, index=0)

def aplicar_filtros(df, col_setor="Setor"):
    out = df.copy()
    if setor_sel != "Todos" and col_setor in out.columns:
        out = out[out[col_setor].astype(str) == setor_sel]
    if mes_sel != "Todos" and "AnoMes" in out.columns:
        out = out[out["AnoMes"].astype(str) == mes_sel]
    return out

abs_f = aplicar_filtros(abs_df)
turn_f = aplicar_filtros(turn_df)
nov_f = aplicar_filtros(nov_df)
base_f = df_base.copy()
if setor_sel != "Todos":
    base_f = base_f[base_f["SETOR"].astype(str) == setor_sel]
total_ref = len(base_f) if len(base_f) else base_total

logo = Path(__file__).parent / "logo.png"
if logo.exists():
    st.image(str(logo), width=190)
st.markdown("## RH • Pessoas")
st.caption("Controle de absenteísmo, turnover, novas contratações e indicadores automáticos por setor e mês.")

c1, c2, c3, c4 = st.columns(4)
with c1:
    kpi_card("Colaboradores Base", total_ref, "Quadro de referência do filtro")
with c2:
    kpi_card("Absenteísmo", len(abs_f), f"Taxa: {taxa_abs:.2f}% no mês atual")
with c3:
    kpi_card("Turnover", len(turn_f), f"Taxa: {taxa_turn:.2f}% no mês atual")
with c4:
    kpi_card("Novas Contratações", len(nov_f), "Lançamentos no filtro atual")

t1, t2, t3, t4 = st.tabs(["Resumo RH", "Absenteísmo", "Turnover", "Novas Contratações"])

with t1:
    r1, r2 = st.columns([1.1, 0.9])
    with r1:
        st.subheader("Quadro atual de colaboradores")
        st.dataframe(base_f.rename(columns={"SETOR": "Setor"}), use_container_width=True, height=450)

        resumo_setor = base_f.groupby("SETOR", dropna=False).size().reset_index(name="Colaboradores").rename(columns={"SETOR": "Setor"}).sort_values("Colaboradores", ascending=False)
        if not resumo_setor.empty:
            plot_bar(resumo_setor, "Setor", "Colaboradores", "Colaboradores por setor")

    with r2:
        st.subheader("Indicadores mensais")
        meses_abs = abs_df.groupby("MesLabel", dropna=False).size().reset_index(name="Absenteísmo")
        meses_turn = turn_df.groupby("MesLabel", dropna=False).size().reset_index(name="Turnover")
        meses_nov = nov_df.groupby("MesLabel", dropna=False).size().reset_index(name="Contratações")

        if not meses_abs.empty:
            plot_bar(meses_abs, "MesLabel", "Absenteísmo", "Absenteísmo por mês")
        if not meses_turn.empty:
            plot_bar(meses_turn, "MesLabel", "Turnover", "Turnover por mês")
        if not meses_nov.empty:
            plot_bar(meses_nov, "MesLabel", "Contratações", "Contratações por mês")


with t2:
    st.subheader("Lançamento de Absenteísmo")
    c1, c2 = st.columns(2)
    with c1:
        data_abs = st.date_input("Data da falta", value=date.today(), key="data_abs")
        nome_abs = st.selectbox("Nome do colaborador", options=nomes, key="nome_abs")
        motivo_abs = st.selectbox("Motivo", ["Falta", "Atestado", "Atraso", "Licença", "Outro"], key="motivo_abs")
    info = info_colaborador(df_base, nome_abs)
    with c2:
        st.text_input("Cadastro", value=str(info["Cadastro"]), disabled=True, key="cad_abs_view")
        st.text_input("Cargo", value=str(info["Cargo"]), disabled=True, key="cargo_abs_view")
        st.text_input("Setor", value=str(info["SETOR"]), disabled=True, key="setor_abs_view")
    obs_abs = st.text_area("Observação", key="obs_abs")
    if st.button("Salvar absenteísmo", use_container_width=True, key="btn_abs"):
        row = [
            data_abs.strftime("%d/%m/%Y"),
            info["Cadastro"],
            info["Nome"],
            info["Admissão"],
            info["Cargo"],
            info["SETOR"],
            motivo_abs,
            obs_abs,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        ]
        append_row("ABSENTEISMO", row)
        st.success("Absenteísmo lançado com sucesso.")
        st.rerun()

    st.markdown("### Histórico de Absenteísmo")
    if not abs_f.empty:
        resumo_abs_setor = abs_f.groupby("Setor", dropna=False).size().reset_index(name="Ocorrências").sort_values("Ocorrências", ascending=False)
        c1, c2 = st.columns([0.9, 1.1])
        with c1:
            plot_bar(resumo_abs_setor, "Setor", "Ocorrências", "Absenteísmo por setor")
        with c2:
            st.dataframe(abs_f, use_container_width=True, height=420)
    else:
        st.info("Sem registros de absenteísmo no filtro atual.")

with t3:
    st.subheader("Lançamento de Turnover / Desligamento")
    c1, c2 = st.columns(2)
    with c1:
        data_turn = st.date_input("Data do desligamento", value=date.today(), key="data_turn")
        nome_turn = st.selectbox("Nome do colaborador desligado", options=nomes, key="nome_turn")
        motivo_turn = st.selectbox("Motivo", ["Pedido de demissão", "Desligamento empresa", "Término contrato", "Outro"], key="motivo_turn")
    info_t = info_colaborador(df_base, nome_turn)
    with c2:
        st.text_input("Cadastro ", value=str(info_t["Cadastro"]), disabled=True, key="cad_turn_view")
        st.text_input("Cargo ", value=str(info_t["Cargo"]), disabled=True, key="cargo_turn_view")
        st.text_input("Setor ", value=str(info_t["SETOR"]), disabled=True, key="setor_turn_view")
    obs_turn = st.text_area("Observação ", key="obs_turn")
    if st.button("Salvar turnover", use_container_width=True, key="btn_turn"):
        row = [
            data_turn.strftime("%d/%m/%Y"),
            info_t["Cadastro"],
            info_t["Nome"],
            info_t["Admissão"],
            info_t["Cargo"],
            info_t["SETOR"],
            motivo_turn,
            obs_turn,
            datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        ]
        append_row("TURNOVER", row)
        st.success("Turnover lançado com sucesso.")
        st.rerun()

    st.markdown("### Histórico de Turnover")
    if not turn_f.empty:
        resumo_turn_setor = turn_f.groupby("Setor", dropna=False).size().reset_index(name="Desligamentos").sort_values("Desligamentos", ascending=False)
        c1, c2 = st.columns([0.9, 1.1])
        with c1:
            plot_bar(resumo_turn_setor, "Setor", "Desligamentos", "Turnover por setor")
        with c2:
            st.dataframe(turn_f, use_container_width=True, height=420)
    else:
        st.info("Sem registros de turnover no filtro atual.")


with t4:
    st.subheader("Lançamento de Novas Contratações")
    with st.form("form_nova_contratacao", clear_on_submit=False):
        c1, c2 = st.columns(2)
        with c1:
            data_reg = st.date_input("Data do registro", value=date.today(), key="data_reg")
            cadastro = st.text_input("Cadastro", key="cad_novo")
            nome = st.text_input("Nome", key="nome_novo")
            admissao = st.date_input("Data de admissão", value=date.today(), key="adm_nova")
        with c2:
            cargo = st.text_input("Cargo", key="cargo_novo")
            setor = st.text_input("Setor", key="setor_novo")
            origem = st.selectbox("Origem", ["Contratação nova", "Recontratação", "Transferência"], key="origem_nova")
        obs_nova = st.text_area("Observação", key="obs_nova")
        enviar_nova = st.form_submit_button("Salvar contratação", use_container_width=True)
        if enviar_nova:
            row = [
                data_reg.strftime("%d/%m/%Y"),
                cadastro,
                nome,
                admissao.strftime("%d/%m/%Y"),
                cargo,
                setor,
                origem,
                obs_nova,
                datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            ]
            append_row("NOVAS_CONTRATACOES", row)
            st.success("Nova contratação lançada com sucesso.")
            st.rerun()

    st.markdown("### Histórico de Novas Contratações")
    if not nov_f.empty:
        resumo_nov_setor = nov_f.groupby("Setor", dropna=False).size().reset_index(name="Contratações").sort_values("Contratações", ascending=False)
        c1, c2 = st.columns([0.9, 1.1])
        with c1:
            plot_bar(resumo_nov_setor, "Setor", "Contratações", "Contratações por setor")
        with c2:
            st.dataframe(nov_f, use_container_width=True, height=420)
    else:
        st.info("Sem registros de novas contratações no filtro atual.")
