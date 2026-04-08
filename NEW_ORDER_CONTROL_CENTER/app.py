
import os
import runpy
from pathlib import Path
from datetime import datetime
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import plotly.graph_objects as go

BASE_DIR = Path(__file__).parent
LEGACY_DIR = BASE_DIR / "legacy_apps"
CFG_PATH = BASE_DIR / "config_mes.xlsx"

st.set_page_config(page_title="NEW ORDER CONTROL CENTER", page_icon="🏭", layout="wide", initial_sidebar_state="expanded")

def inject_theme():
    st.markdown("""
    <style>
    .stApp {
        background:
            radial-gradient(circle at 12% 10%, rgba(0,195,255,0.08), transparent 18%),
            radial-gradient(circle at 88% 14%, rgba(255,120,0,0.08), transparent 18%),
            linear-gradient(135deg, #071018 0%, #0b131d 40%, #0b1118 100%);
        color: #edf3f9;
    }
    section[data-testid="stSidebar"] {
        background:
            linear-gradient(180deg, rgba(255,255,255,0.04), rgba(255,255,255,0.015)),
            linear-gradient(180deg, #09111a 0%, #111a25 100%);
        border-right: 1px solid rgba(255,255,255,0.08);
    }
    .block-container { padding-top: 1rem; padding-bottom: 2rem; max-width: 1450px; }
    .hero, .kpi-card, .info-card, .login-card {
        border-radius: 24px;
        border: 1px solid rgba(255,255,255,0.08);
        background: linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.02));
        box-shadow: 0 14px 34px rgba(0,0,0,0.24);
    }
    .hero { padding: 32px 34px; margin-bottom: 20px; }
    .kpi-card { padding: 18px; min-height: 165px; }
    .info-card { padding: 18px; min-height: 220px; }
    .login-card { padding: 24px; max-width: 460px; margin: 40px auto; }
    .title { font-size: 48px; font-weight: 950; line-height: 0.98; color: #fff; margin-bottom: 12px; }
    .subtitle { color: #b9c5d2; font-size: 14px; text-transform: uppercase; letter-spacing: 1.6px; margin-bottom: 16px; }
    .line { width: 220px; height: 3px; border-radius: 999px; background: linear-gradient(90deg, #00c3ff, #ff7800); margin-bottom: 18px; }
    .hero-text { color: #d8e1ea; font-size: 15px; line-height: 1.65; max-width: 880px; margin-bottom: 18px; }
    .badge { display: inline-flex; align-items: center; gap: 8px; padding: 8px 12px; border-radius: 999px; background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.08); color: #e8f0f7; font-size: 12px; font-weight: 700; margin: 6px 8px 0 0; }
    .section-title { color: #fff; font-size: 24px; font-weight: 900; margin: 10px 0 14px 0; }
    .module-card { position: relative; overflow: hidden; border-radius: 24px; padding: 22px; min-height: 210px; background: linear-gradient(180deg, rgba(255,255,255,0.06), rgba(255,255,255,0.02)), linear-gradient(135deg, rgba(0,195,255,0.03), rgba(255,120,0,0.03)); border: 1px solid rgba(255,255,255,0.08); box-shadow: 0 14px 34px rgba(0,0,0,0.25), inset 0 1px 0 rgba(255,255,255,0.06); }
    .module-card:before { content:""; position:absolute; left:0; top:0; width:100%; height:4px; background: linear-gradient(90deg, #00c3ff, #ff7800); opacity:0.95; }
    .module-icon { font-size:34px; margin-bottom:8px; }
    .module-tag { display:inline-block; padding:6px 10px; border-radius:999px; font-size:11px; font-weight:800; text-transform: uppercase; letter-spacing:1px; color:#dce7f2; background: rgba(255,255,255,0.06); border: 1px solid rgba(255,255,255,0.08); margin-bottom:14px; }
    .module-title { color:#fff; font-size:25px; font-weight:900; margin-bottom:10px; }
    .module-desc { color:#a9b8c8; font-size:13px; line-height:1.55; }
    div.stButton > button { width: 100%; min-height: 76px !important; border-radius: 18px !important; font-size: 18px !important; font-weight: 900 !important; border: 1px solid rgba(255,255,255,0.10) !important; background: linear-gradient(135deg, rgba(0,195,255,0.14), rgba(255,120,0,0.10)) !important; color: #ffffff !important; box-shadow: 0 14px 28px rgba(0,0,0,0.22); }
    div.stButton > button:hover { border-color: rgba(255,255,255,0.18) !important; box-shadow: 0 18px 34px rgba(0,0,0,0.28), 0 0 16px rgba(0,195,255,0.10); }
    .kpi-label { color: #9fb0c2; font-size: 12px; text-transform: uppercase; font-weight: 800; letter-spacing: 1px; margin-bottom: 8px; }
    .kpi-value { color: #ffffff; font-size: 34px; font-weight: 950; line-height: 1.0; margin-bottom: 8px; }
    .kpi-sub { color: #9aabbd; font-size: 12px; line-height: 1.45; }
    .sem-pill { display:inline-flex; align-items:center; gap:8px; padding:6px 10px; border-radius:999px; margin-top:10px; font-size:11px; font-weight:800; text-transform:uppercase; letter-spacing:1px; border:1px solid rgba(255,255,255,0.10); background: rgba(255,255,255,0.05); color:#edf3f9; }
    .sem-dot { width:10px; height:10px; border-radius:50%; display:inline-block; box-shadow: 0 0 10px currentColor; }
    .green { background:#21d07a; color:#21d07a; } .orange { background:#ffb020; color:#ffb020; } .red { background:#ff5a5f; color:#ff5a5f; }
    .info-title { color: #ffffff; font-size: 20px; font-weight: 900; margin-bottom: 8px; }
    .info-item { border-left: 3px solid rgba(0,195,255,0.55); padding-left: 14px; margin: 14px 0; }
    .info-head { color: #ffffff; font-size: 15px; font-weight: 800; }
    .info-text { color: #9aabbd; font-size: 12px; margin-top: 4px; line-height: 1.45; }
    </style>
    """, unsafe_allow_html=True)

def run_legacy_app(path: Path):
    original_cwd = Path.cwd()
    original_fn = st.set_page_config
    try:
        os.chdir(path.parent)
        st.set_page_config = lambda *args, **kwargs: None
        runpy.run_path(str(path), run_name="__main__")
    finally:
        st.set_page_config = original_fn
        os.chdir(original_cwd)

def read_excel_safe(path: Path, **kwargs):
    try:
        return pd.read_excel(path, engine="openpyxl", **kwargs)
    except Exception:
        return pd.DataFrame()

def find_col(df, keywords):
    if df is None or df.empty:
        return None
    cols = [str(c).strip() for c in df.columns]
    lower_map = {str(c).strip().lower(): c for c in df.columns}
    for key in keywords:
        for c in cols:
            if key.lower() == c.lower():
                return lower_map[c.lower()]
    for key in keywords:
        for c in cols:
            if key.lower() in c.lower():
                return lower_map[c.lower()]
    return None

def parse_date_series(s):
    try:
        return pd.to_datetime(s, errors="coerce", dayfirst=True)
    except Exception:
        return pd.to_datetime(pd.Series([], dtype="object"))

def semaforo_html(status):
    status = str(status).lower()
    if status == "bom":
        return '<span class="sem-pill"><span class="sem-dot green"></span>Bom</span>'
    if status in ("atenção","atencao"):
        return '<span class="sem-pill"><span class="sem-dot orange"></span>Atenção</span>'
    return '<span class="sem-pill"><span class="sem-dot red"></span>Crítico</span>'

def load_users():
    df = read_excel_safe(CFG_PATH, sheet_name="USUARIOS")
    if df.empty:
        return pd.DataFrame(columns=["usuario","senha","perfil","nome"])
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

def load_metas():
    df = read_excel_safe(CFG_PATH, sheet_name="METAS")
    metas = {}
    if df.empty:
        return metas
    df.columns = [str(c).strip().lower() for c in df.columns]
    for _, row in df.iterrows():
        metas[str(row.get("indicador","")).upper()] = {
            "bom": float(row.get("meta_bom", 0) or 0),
            "atencao": float(row.get("meta_atencao", 0) or 0),
            "obs": str(row.get("observacao",""))
        }
    return metas

def append_row(sheet_name, values):
    wb = load_workbook(CFG_PATH)
    ws = wb[sheet_name]
    ws.append(values)
    wb.save(CFG_PATH)

def ensure_history_snapshot(snapshot):
    df = read_excel_safe(CFG_PATH, sheet_name="HISTORICO_MENSAL")
    if not df.empty:
        cols = [str(c).strip().lower() for c in df.columns]
        if "ano_mes" in cols:
            col = df.columns[cols.index("ano_mes")]
            if snapshot["ano_mes"] in df[col].astype(str).tolist():
                return
    append_row("HISTORICO_MENSAL", [
        snapshot["ano_mes"], snapshot["produzido_mes"], snapshot["forecast_mes"], snapshot["faturado_mes"],
        snapshot["gargalo"], snapshot["carga_critica"], snapshot["acoes_abertas"], snapshot["acoes_atrasadas"],
        snapshot["efetivo"], snapshot["abs_mes"], snapshot["turn_mes"],
        snapshot["status_producao"], snapshot["status_carga"], snapshot["status_pa"], snapshot["status_rh"]
    ])

def append_alerts_if_needed(items):
    df = read_excel_safe(CFG_PATH, sheet_name="ALERTAS")
    existing = set()
    if not df.empty:
        cols = [str(c).strip().lower() for c in df.columns]
        if "modulo" in cols and "mensagem" in cols:
            mod_col = df.columns[cols.index("modulo")]
            msg_col = df.columns[cols.index("mensagem")]
            existing = set(zip(df[mod_col].astype(str), df[msg_col].astype(str)))
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for modulo, status, msg in items:
        if status in ("crítico","critico","atenção","atencao") and (modulo, msg) not in existing:
            append_row("ALERTAS", [now, modulo, status, msg])


def get_producao_kpis(metas):
    path = LEGACY_DIR / "producao" / "PROD-PRODT.xlsx"
    if not path.exists():
        return {"produzido_mes":0,"faturado_mes":0,"forecast_mes":0,"status":"crítico"}

    hoje = pd.Timestamp.today()
    mes = int(hoje.month)

    fat_alias = {
        1: ["FAT.JAN", "FAT JAN", "FAT_JAN", "JAN"],
        2: ["FAT.FEV", "FAT FEV", "FAT_FEV", "FEV"],
        3: ["FAT.MAR", "FAT MAR", "FAT_MAR", "MAR"],
        4: ["FAT.ABR", "FAT.ABRIL", "FAT ABR", "FAT ABRIL", "ABR", "ABRIL"],
        5: ["FAT.MAI", "FAT MAI", "FAT_MAI", "MAI"],
        6: ["FAT.JUN", "FAT JUN", "FAT_JUN", "JUN"],
        7: ["FAT.JUL", "FAT JUL", "FAT_JUL", "JUL"],
        8: ["FAT.AGO", "FAT AGO", "FAT_AGO", "AGO"],
        9: ["FAT.SET", "FAT SET", "FAT_SET", "SET"],
        10: ["FAT.OUT", "FAT OUT", "FAT_OUT", "OUT"],
        11: ["FAT.NOV", "FAT NOV", "FAT_NOV", "NOV"],
        12: ["FAT.DEZ", "FAT DEZ", "FAT_DEZ", "DEZ"],
    }
    for_alias = {
        1: ["FOR. JAN", "FOR.JAN", "FOR JAN", "JAN"],
        2: ["FOR. FEV", "FOR.FEV", "FOR FEV", "FEV"],
        3: ["FOR. MAR", "FOR.MAR", "FOR MAR", "MAR"],
        4: ["FOR. ABR", "FOR.ABR", "FOR ABR", "FOR. ABRIL", "FOR.ABRIL", "ABR", "ABRIL"],
        5: ["FOR. MAI", "FOR.MAI", "FOR MAI", "MAI"],
        6: ["FOR. JUN", "FOR.JUN", "FOR JUN", "JUN"],
        7: ["FOR. JUL", "FOR.JUL", "FOR JUL", "JUL"],
        8: ["FOR. AGO", "FOR.AGO", "FOR AGO", "AGO"],
        9: ["FOR. SET", "FOR.SET", "FOR SET", "SET"],
        10: ["FOR. OUT", "FOR.OUT", "FOR OUT", "OUT"],
        11: ["FOR. NOV", "FOR.NOV", "FOR NOV", "NOV"],
        12: ["FOR. DEZ", "FOR.DEZ", "FOR DEZ", "DEZ"],
    }

    produzido_total = 0.0
    faturado_total = 0.0
    forecast_total = 0.0

    try:
        xl = pd.ExcelFile(path, engine="openpyxl")
        sheets = xl.sheet_names
    except Exception:
        sheets = []

    for sheet in sheets[:5]:
        try:
            df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        except Exception:
            continue
        prod_col = find_col(df, ["produzido", "produção", "qtde produzida", "qtd produzida"])
        if prod_col and prod_col in df.columns:
            produzido_total += pd.to_numeric(df[prod_col], errors="coerce").fillna(0).sum()

    if "FATURADO" in sheets:
        fat_df = read_excel_safe(path, sheet_name="FATURADO")
        if not fat_df.empty:
            fat_df.columns = [str(c).strip().upper() for c in fat_df.columns]
            fat_col = None
            for alias in fat_alias.get(mes, []):
                if alias.upper() in fat_df.columns:
                    fat_col = alias.upper()
                    break
            if fat_col is None:
                for c in fat_df.columns:
                    norm_c = c.replace(".", "").replace(" ", "")
                    for a in fat_alias.get(mes, []):
                        if a.replace(".", "").replace(" ", "") in norm_c:
                            fat_col = c
                            break
                    if fat_col:
                        break
            if fat_col:
                faturado_total = pd.to_numeric(fat_df[fat_col], errors="coerce").fillna(0).sum()
            elif "FAT.TOTAL" in fat_df.columns:
                faturado_total = pd.to_numeric(fat_df["FAT.TOTAL"], errors="coerce").fillna(0).sum()

    if "FORECAST" in sheets:
        for_df = read_excel_safe(path, sheet_name="FORECAST")
        if not for_df.empty:
            for_df.columns = [str(c).strip().upper() for c in for_df.columns]
            for_col = None
            for alias in for_alias.get(mes, []):
                if alias.upper() in for_df.columns:
                    for_col = alias.upper()
                    break
            if for_col is None:
                for c in for_df.columns:
                    norm_c = c.replace(".", "").replace(" ", "")
                    for a in for_alias.get(mes, []):
                        if a.replace(".", "").replace(" ", "") in norm_c:
                            for_col = c
                            break
                    if for_col:
                        break
            if for_col:
                forecast_total = pd.to_numeric(for_df[for_col], errors="coerce").fillna(0).sum()

    m = metas.get("PRODUCAO_RATIO", {"bom":1.0,"atencao":0.85})
    ratio = (produzido_total / forecast_total) if forecast_total > 0 else 0
    status = "bom" if ratio >= m["bom"] else ("atenção" if ratio >= m["atencao"] else "crítico")

    return {
        "produzido_mes": round(float(produzido_total), 0),
        "faturado_mes": round(float(faturado_total), 0),
        "forecast_mes": round(float(forecast_total), 0),
        "status": status
    }


def get_carga_kpis(metas):
    path = LEGACY_DIR / "carga_maquina" / "CG BOT PY.xlsx"
    if not path.exists():
        return {"gargalo":"-","utilizacao":0,"linhas":0,"status":"crítico"}
    df = read_excel_safe(path)
    if df.empty:
        return {"gargalo":"-","utilizacao":0,"linhas":0,"status":"crítico"}
    tempo_col = find_col(df, ["tempo individual","tempo","min"])
    desc_col = df.columns[5] if df.shape[1] > 5 else None
    cr_col = find_col(df, ["cr"])
    tempos = pd.to_numeric(df[tempo_col], errors="coerce").fillna(0) if tempo_col else pd.Series([0]*len(df))
    if desc_col:
        agg = df.assign(_tempo=tempos).groupby(desc_col, dropna=False)["_tempo"].sum().sort_values(ascending=False)
        gargalo = str(agg.index[0]) if len(agg) else "-"
        util = float(agg.iloc[0]) if len(agg) else 0
    else:
        gargalo, util = "-", 0
    linhas = int(df[cr_col].nunique()) if cr_col and cr_col in df.columns else len(df)
    m = metas.get("CARGA_CRITICA", {"bom":5000,"atencao":12000})
    status = "bom" if util <= m["bom"] else ("atenção" if util <= m["atencao"] else "crítico")
    return {"gargalo":gargalo,"utilizacao":round(util,0),"linhas":linhas,"status":status}

def get_rh_kpis(metas):
    path = LEGACY_DIR / "rh" / "QUADRO COLABORADORES.xlsx"
    if not path.exists():
        return {"efetivo":0,"abs_mes":0,"turn_mes":0,"status":"crítico"}
    base = read_excel_safe(path, sheet_name="RELAÇÃO DE COLABORADORES 19")
    abs_df = read_excel_safe(path, sheet_name="ABSENTEISMO")
    turn_df = read_excel_safe(path, sheet_name="TURNOVER")
    efetivo = len(base) if not base.empty else 0
    hoje = pd.Timestamp.today()
    def conta_mes(df):
        if df.empty:
            return 0
        col = find_col(df, ["data"])
        if not col:
            return 0
        dt = parse_date_series(df[col])
        return int(((dt.dt.month == hoje.month) & (dt.dt.year == hoje.year)).sum())
    abs_mes, turn_mes = conta_mes(abs_df), conta_mes(turn_df)
    rate = max((abs_mes/efetivo*100) if efetivo else 0, (turn_mes/efetivo*100) if efetivo else 0)
    m = metas.get("RH_TAXA", {"bom":2,"atencao":5})
    status = "bom" if rate <= m["bom"] else ("atenção" if rate <= m["atencao"] else "crítico")
    return {"efetivo":efetivo,"abs_mes":abs_mes,"turn_mes":turn_mes,"status":status}

def get_pa_kpis(metas):
    path = LEGACY_DIR / "pa" / "data" / "pa.xlsx"
    if not path.exists():
        path = LEGACY_DIR / "pa" / "pa.xlsx"
    if not path.exists():
        return {"abertas":0,"atrasadas":0,"responsavel":"-","status":"crítico"}
    try:
        raw = pd.read_excel(path, sheet_name="PA", engine="openpyxl", header=None).dropna(how="all")
    except Exception:
        return {"abertas":0,"atrasadas":0,"responsavel":"-","status":"crítico"}
    header_row = None
    for i, row in raw.iterrows():
        vals = row.astype(str)
        if vals.str.contains("Ação", case=False, na=False).any() and vals.str.contains("Status", case=False, na=False).any():
            header_row = i
            break
    if header_row is None:
        return {"abertas":0,"atrasadas":0,"responsavel":"-","status":"crítico"}
    df = pd.read_excel(path, sheet_name="PA", engine="openpyxl", header=header_row).dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    status_col = find_col(df, ["status"])
    prazo_col = find_col(df, ["prazo"])
    resp_col = find_col(df, ["responsável","responsavel"])
    status_s = df[status_col].astype(str).str.strip().str.lower() if status_col and status_col in df.columns else pd.Series([], dtype="object")
    abertas = int(status_s.isin(["aberta","em execução","em execucao","atrasada","open"]).sum()) if not status_s.empty else len(df)
    atrasadas = 0
    if prazo_col and prazo_col in df.columns:
        dt = parse_date_series(df[prazo_col])
        closed = status_s.isin(["executado","cancelada"]) if not status_s.empty else pd.Series([False]*len(df))
        atrasadas = int(((dt < pd.Timestamp.today().normalize()) & (~closed)).sum())
    responsavel = "-"
    if resp_col and resp_col in df.columns:
        top = df[resp_col].astype(str).str.strip()
        top = top[top.ne("") & top.ne("nan")]
        if not top.empty:
            responsavel = top.value_counts().index[0]
    m = metas.get("PA_ATRASADAS", {"bom":0,"atencao":3})
    status = "bom" if atrasadas <= m["bom"] else ("atenção" if atrasadas <= m["atencao"] else "crítico")
    return {"abertas":abertas,"atrasadas":atrasadas,"responsavel":responsavel,"status":status}

def kpi_card(label, value, sub, status):
    st.markdown(f'''
    <div class="kpi-card">
        <div class="kpi-label">{label}</div>
        <div class="kpi-value">{value}</div>
        <div class="kpi-sub">{sub}</div>
        {semaforo_html(status)}
    </div>
    ''', unsafe_allow_html=True)

def module_card(icon, tag, title, desc):
    st.markdown(f'''
    <div class="module-card">
        <div class="module-icon">{icon}</div>
        <div class="module-tag">{tag}</div>
        <div class="module-title">{title}</div>
        <div class="module-desc">{desc}</div>
    </div>
    ''', unsafe_allow_html=True)

def plot_history():
    df = read_excel_safe(CFG_PATH, sheet_name="HISTORICO_MENSAL")
    if df.empty:
        st.info("Sem histórico mensal registrado ainda.")
        return
    df.columns = [str(c).strip().lower() for c in df.columns]
    if "ano_mes" not in df.columns:
        st.info("Histórico mensal sem coluna ano_mes.")
        return
    fig = go.Figure()
    for col, name in [("produzido_mes","Produzido"),("forecast_mes","Forecast"),("faturado_mes","Faturado")]:
        if col in df.columns:
            fig.add_trace(go.Scatter(x=df["ano_mes"], y=pd.to_numeric(df[col], errors="coerce"), mode="lines+markers", name=name))
    fig.update_layout(title="Tendência Mensal", height=360, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(255,255,255,0.02)", font=dict(color="#edf3f9"), margin=dict(l=20, r=20, t=50, b=20), legend=dict(orientation="h", y=1.1, x=0))
    st.plotly_chart(fig, use_container_width=True)

def render_alerts():
    df = read_excel_safe(CFG_PATH, sheet_name="ALERTAS")
    st.markdown('<div class="section-title">Alertas Automáticos</div>', unsafe_allow_html=True)
    if df.empty:
        st.success("Sem alertas automáticos registrados.")
        return
    st.dataframe(df.tail(20), use_container_width=True, height=260)

def render_config_metas():
    st.markdown('<div class="section-title">Configuração de Metas</div>', unsafe_allow_html=True)
    metas_df = read_excel_safe(CFG_PATH, sheet_name="METAS")
    edited = st.data_editor(metas_df, use_container_width=True, num_rows="fixed", key="metas_editor")
    if st.button("Salvar metas", key="salvar_metas"):
        wb = load_workbook(CFG_PATH)
        ws = wb["METAS"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for _, row in edited.iterrows():
            ws.append(list(row.values))
        wb.save(CFG_PATH)
        st.success("Metas salvas com sucesso.")
        st.rerun()

def render_login():
    st.markdown('<div class="login-card">', unsafe_allow_html=True)
    st.markdown("## Login por setor")
    st.caption("Perfis padrão no Excel: diretoria, producao, rh")
    with st.form("login_form"):
        usuario = st.text_input("Usuário")
        senha = st.text_input("Senha", type="password")
        entrar = st.form_submit_button("Entrar", use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)
    if entrar:
        users = load_users()
        if users.empty:
            st.error("Base de usuários não encontrada.")
            return
        match = users[(users["usuario"].astype(str) == usuario) & (users["senha"].astype(str) == senha)]
        if match.empty:
            st.error("Usuário ou senha inválidos.")
        else:
            row = match.iloc[0]
            st.session_state["auth"] = True
            st.session_state["perfil"] = str(row.get("perfil","diretoria"))
            st.session_state["nome"] = str(row.get("nome", usuario))
            st.rerun()

def allowed_options(perfil):
    p = str(perfil).lower()
    if p == "diretoria":
        return ["Home","Produção","Plano de Ação Fábrica","Carga Máquina","RH • Pessoas","MES • Gestão"]
    if p == "producao":
        return ["Home","Produção","Plano de Ação Fábrica","Carga Máquina","MES • Gestão"]
    if p == "rh":
        return ["Home","RH • Pessoas","MES • Gestão"]
    return ["Home"]

def home():
    logo_path = LEGACY_DIR / "producao" / "logo.png"
    if logo_path.exists():
        st.image(str(logo_path), width=210)

    metas = load_metas()
    prod = get_producao_kpis(metas)
    carga = get_carga_kpis(metas)
    rh = get_rh_kpis(metas)
    pa = get_pa_kpis(metas)

    now = pd.Timestamp.today()
    snapshot = {
        "ano_mes": now.strftime("%Y-%m"),
        "produzido_mes": prod["produzido_mes"],
        "forecast_mes": prod["forecast_mes"],
        "faturado_mes": prod["faturado_mes"],
        "gargalo": carga["gargalo"],
        "carga_critica": carga["utilizacao"],
        "acoes_abertas": pa["abertas"],
        "acoes_atrasadas": pa["atrasadas"],
        "efetivo": rh["efetivo"],
        "abs_mes": rh["abs_mes"],
        "turn_mes": rh["turn_mes"],
        "status_producao": prod["status"],
        "status_carga": carga["status"],
        "status_pa": pa["status"],
        "status_rh": rh["status"],
    }
    ensure_history_snapshot(snapshot)
    append_alerts_if_needed([
        ("Produção", prod["status"], f"Produção em {prod['status']} - Produzido {prod['produzido_mes']} x Forecast {prod['forecast_mes']}"),
        ("Carga Máquina", carga["status"], f"Carga em {carga['status']} - Gargalo {carga['gargalo']}"),
        ("Plano de Ação", pa["status"], f"Plano de Ação em {pa['status']} - Atrasadas {pa['atrasadas']}"),
        ("RH", rh["status"], f"RH em {rh['status']} - Abs {rh['abs_mes']} / Turn {rh['turn_mes']}"),
    ])

    st.markdown('''
    <div class="hero">
        <div class="title">NEW ORDER<br>CONTROL CENTER</div>
        <div class="subtitle">Central industrial integrada • nível MES • visão executiva de fábrica</div>
        <div class="line"></div>
        <div class="hero-text">
            Plataforma corporativa com metas configuráveis, alertas automáticos, histórico mensal e acesso por perfil.
        </div>
        <span class="badge">🔵 Produção</span>
        <span class="badge">🟠 Planejamento</span>
        <span class="badge">🟢 Pessoas</span>
        <span class="badge">🔴 Alertas operacionais</span>
    </div>
    ''', unsafe_allow_html=True)

    st.markdown('<div class="section-title">Módulos Principais</div>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        if st.button("🏭 PRODUÇÃO", use_container_width=True):
            st.session_state["menu_target"] = "Produção"
            st.rerun()
    with c2:
        if st.button("📋 PLANO DE AÇÃO", use_container_width=True):
            st.session_state["menu_target"] = "Plano de Ação Fábrica"
            st.rerun()
    with c3:
        if st.button("⚙️ CARGA MÁQUINA", use_container_width=True):
            st.session_state["menu_target"] = "Carga Máquina"
            st.rerun()
    with c4:
        if st.button("👥 RH • PESSOAS", use_container_width=True):
            st.session_state["menu_target"] = "RH • Pessoas"
            st.rerun()

    st.markdown('<div class="section-title">KPIs Executivos da Home</div>', unsafe_allow_html=True)
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        kpi_card("Produzido mês", f"{prod['produzido_mes']:,.0f}".replace(",", "."), f"Forecast: {prod['forecast_mes']:,.0f}".replace(",", ".") + f" • Faturado: {prod['faturado_mes']:,.0f}".replace(",", "."), prod["status"])
    with k2:
        kpi_card("Gargalo atual", str(carga["gargalo"]), f"Carga crítica: {carga['utilizacao']:,.0f}".replace(",", ".") + f" • Linhas/CR: {carga['linhas']}", carga["status"])
    with k3:
        kpi_card("Plano de ação", str(pa["abertas"]), f"Atrasadas: {pa['atrasadas']} • Resp.: {pa['responsavel']}", pa["status"])
    with k4:
        kpi_card("Efetivo RH", str(rh["efetivo"]), f"Absenteísmo mês: {rh['abs_mes']} • Turnover mês: {rh['turn_mes']}", rh["status"])

    st.markdown('<div class="section-title">Visão dos Módulos</div>', unsafe_allow_html=True)
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        module_card("🏭", "Operação", "Produção", "Dashboard produtivo com produtividade, tendência mensal, forecast x produzido x faturado, rankings e leitura direta da base operacional.")
    with m2:
        module_card("📋", "Gestão à vista", "Plano de Ação Fábrica", "Gemba Board ANDON com KPIs, Pareto, Kanban, saúde do plano de ação, atrasos críticos e acompanhamento por responsável e setor.")
    with m3:
        module_card("⚙️", "Capacidade", "Carga Máquina", "Simulação industrial com gargalo matemático real, mão de obra, indiretos, TAKT, previsão de cenário e análise executiva de utilização.")
    with m4:
        module_card("👥", "Pessoas", "RH • Pessoas", "Controle de absenteísmo, turnover e novas contratações com lançamento diário, indicadores mensais e análise por setor.")

def render_mes_gestao():
    metas = load_metas()
    prod = get_producao_kpis(metas)
    carga = get_carga_kpis(metas)
    rh = get_rh_kpis(metas)
    pa = get_pa_kpis(metas)

    st.markdown('<div class="section-title">MES • Gestão</div>', unsafe_allow_html=True)
    a, b, c, d = st.columns(4)
    with a:
        kpi_card("Produção", f"{prod['produzido_mes']:,.0f}".replace(",", "."), f"Forecast {prod['forecast_mes']:,.0f}".replace(",", "."), prod["status"])
    with b:
        kpi_card("Carga", str(carga["gargalo"]), f"Carga crítica {carga['utilizacao']:,.0f}".replace(",", "."), carga["status"])
    with c:
        kpi_card("Plano de ação", str(pa["abertas"]), f"Atrasadas {pa['atrasadas']}", pa["status"])
    with d:
        kpi_card("RH", str(rh["efetivo"]), f"Abs {rh['abs_mes']} • Turn {rh['turn_mes']}", rh["status"])

    left, right = st.columns([1.15, 0.85], gap="large")
    with left:
        plot_history()
    with right:
        render_alerts()

    render_config_metas()

inject_theme()

if "auth" not in st.session_state:
    st.session_state["auth"] = False

if not st.session_state["auth"]:
    render_login()
    st.stop()

perfil = st.session_state.get("perfil", "diretoria")
nome = st.session_state.get("nome", "Usuário")

logo_sidebar = LEGACY_DIR / "producao" / "logo.png"
if logo_sidebar.exists():
    st.sidebar.image(str(logo_sidebar), width=150)
st.sidebar.markdown(f"## GNO • {nome}")
st.sidebar.caption(f"Perfil: {perfil.title()}")
if st.sidebar.button("Sair", use_container_width=True):
    st.session_state.clear()
    st.rerun()

options = allowed_options(perfil)
target = st.session_state.pop("menu_target", None)
default_idx = options.index(target) if target in options else 0
menu = st.sidebar.radio("Navegação", options, index=default_idx)

if menu == "Home":
    home()
elif menu == "Produção":
    run_legacy_app(LEGACY_DIR / "producao" / "app.py")
elif menu == "Plano de Ação Fábrica":
    run_legacy_app(LEGACY_DIR / "pa" / "app.py")
elif menu == "Carga Máquina":
    run_legacy_app(LEGACY_DIR / "carga_maquina" / "app.py")
elif menu == "RH • Pessoas":
    run_legacy_app(LEGACY_DIR / "rh" / "app.py")
elif menu == "MES • Gestão":
    render_mes_gestao()
